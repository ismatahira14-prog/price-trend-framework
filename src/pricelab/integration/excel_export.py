"""Write a human-friendly .xlsx version of master_long for opening in Excel.

The .parquet/.csv outputs stay the machine-readable source of truth; this is
purely a readability convenience: bold header, frozen header row, sane column
widths, an autofilter, and a real date format instead of a raw ISO string.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_WIDTHS = {
    "date": 12,
    "freq": 6,
    "region": 20,
    "region_level": 13,
    "commodity": 34,
    "variable": 16,
    "value": 12,
    "unit": 20,
    "source": 20,
    "is_imputed": 11,
}


def write_excel(df: pd.DataFrame, path: str | Path) -> Path:
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        df.to_excel(writer, sheet_name="master_long", index=False)
        ws = writer.sheets["master_long"]

        for col in ws[1]:
            col.font = Font(bold=True)

        for i, name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = _WIDTHS.get(name, 16)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return path
